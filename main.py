import requests
from bs4 import BeautifulSoup  as s
import json
import openpyxl
#import urllib.parse
import os.path
import telebot
#telgram bot detail
TOKEN = "Your TOKEN"
bot = telebot.TeleBot(TOKEN, parse_mode=None, threaded=True)
# Define the range of rows to process
start_row = 1
end_row = 75000

# Load the last processed row from a file
try:
    with open("last_processed_row.txt", "r") as f:
        last_processed_row = int(f.read().strip())
except FileNotFoundError:
    last_processed_row = start_row

# Load the Excel spreadsheet
wb = openpyxl.load_workbook('Your Excel File')
print(wb.sheetnames)
sheet = wb['Sheet1']

# Loop through the URLs in the spreadsheet
for i2, row in enumerate(sheet.iter_rows(min_row=start_row, max_row=end_row, max_col=1, values_only=True), start=start_row):
    if i2 <= last_processed_row:
        continue
    url = row[0]
    print(f"Processing row {i2}: {url}")
    try:
        #craate function download all data though excel
        # Update the last processed row after every iteration
        last_processed_row = i2
        with open("last_processed_row.txt", "w") as f:
            f.write(str(last_processed_row))

    except Exception as e:
        print(f"Error processing row {i2}: {e}")
        with open("error2.txt", "a") as f:
            f.write(str(url) + "\n")
            i = open("error2.txt", 'rb')
            try:
                bot.send_document(-1001534695986, i ,parse_mode='Markdown')
                print(f"Bot Send error file!")
            except Exception as e:
                print(f"Error processing row : {e}")
                continue
        #with open("error_qurl.txt", "a") as f:
            #f.write(str(qurl) + "\n")
        continue
